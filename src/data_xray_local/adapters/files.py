"""Read supported local files into privacy-detector chunks."""

from __future__ import annotations

import csv
import json
import mimetypes
import os
import re
import sys
import zipfile
from contextlib import suppress
from dataclasses import dataclass
from importlib import import_module
from pathlib import Path
from typing import Any, Protocol, cast

from defusedxml import ElementTree
from openpyxl import load_workbook
from PIL import ExifTags, Image, UnidentifiedImageError

from data_xray_local.domain.models import DataCategory, Severity


class MagicModule(Protocol):
    def from_file(self, filename: str, *, mime: bool = False) -> str: ...


def _load_magic_module(platform_name: str = sys.platform) -> MagicModule | None:
    """Load libmagic only on platforms where its native loader is reliable."""
    if platform_name == "win32":
        return None
    try:
        return cast(MagicModule, import_module("magic"))
    except (ImportError, OSError):  # pragma: no cover - environment-specific optional native load
        return None


magic_module = _load_magic_module()


TEXT_SUFFIXES = {
    ".cfg",
    ".conf",
    ".csv",
    ".env",
    ".ini",
    ".json",
    ".log",
    ".md",
    ".rst",
    ".text",
    ".toml",
    ".tsv",
    ".txt",
    ".yaml",
    ".yml",
}
OFFICE_SUFFIXES = {".docx", ".pptx", ".xlsx"}
IMAGE_SUFFIXES = {".jpeg", ".jpg", ".png", ".tif", ".tiff", ".webp"}
SUPPORTED_SUFFIXES = TEXT_SUFFIXES | OFFICE_SUFFIXES | IMAGE_SUFFIXES
DEFAULT_EXCLUDED_DIRECTORIES = {
    ".git",
    ".mypy_cache",
    ".pytest_cache",
    ".ruff_cache",
    ".venv",
    "__pycache__",
    "build",
    "dist",
    "dist-release",
    "htmlcov",
    "node_modules",
    "reports",
}


class FileAccessError(RuntimeError):
    """A safe extraction failure whose message never contains source content."""


class UnsupportedFileError(FileAccessError):
    """File type is outside the documented MVP support."""


class FileTooLargeError(FileAccessError):
    """File was deliberately skipped to bound memory use."""


@dataclass(frozen=True, slots=True)
class ContentChunk:
    location: str
    text: str


@dataclass(frozen=True, slots=True)
class MetadataSignal:
    category: DataCategory
    severity: Severity
    location: str
    value: str
    rule_id: str
    remediation: str


@dataclass(frozen=True, slots=True)
class ExtractedDocument:
    file_type: str
    mime_type: str
    chunks: tuple[ContentChunk, ...]
    metadata_signals: tuple[MetadataSignal, ...] = ()


def _safe_error(error: BaseException) -> str:
    labels = {
        PermissionError: "permission denied",
        UnicodeDecodeError: "text encoding could not be decoded",
        json.JSONDecodeError: "invalid JSON syntax",
        zipfile.BadZipFile: "invalid or damaged Office container",
        ElementTree.ParseError: "invalid Office XML",
        UnidentifiedImageError: "unsupported or damaged image",
    }
    for error_type, label in labels.items():
        if isinstance(error, error_type):
            return label
    return f"{type(error).__name__}: extraction failed"


def _flatten_json(value: Any, location: str = "$") -> list[ContentChunk]:
    chunks: list[ContentChunk] = []
    if isinstance(value, dict):
        for key in sorted(value):
            child = (
                f"{location}.{key}"
                if re.fullmatch(r"[A-Za-z_]\w*", str(key))
                else f"{location}[key]"
            )
            chunks.extend(_flatten_json(value[key], child))
    elif isinstance(value, list):
        for index, item in enumerate(value):
            chunks.extend(_flatten_json(item, f"{location}[{index}]"))
    elif value is not None:
        chunks.append(ContentChunk(location=location, text=str(value)))
    return chunks


def _office_core_chunks(
    archive: zipfile.ZipFile,
) -> tuple[list[ContentChunk], list[MetadataSignal]]:
    chunks: list[ContentChunk] = []
    signals: list[MetadataSignal] = []
    try:
        raw = archive.read("docProps/core.xml")
    except KeyError:
        return chunks, signals
    root = ElementTree.fromstring(raw)
    author_fields = {"creator", "lastModifiedBy"}
    for element in root.iter():
        field_name = element.tag.rsplit("}", 1)[-1]
        value = (element.text or "").strip()
        if not value:
            continue
        chunks.append(ContentChunk(location=f"office metadata · {field_name}", text=value))
        if field_name in author_fields:
            signals.append(
                MetadataSignal(
                    category=DataCategory.AUTHOR_METADATA,
                    severity=Severity.MEDIUM,
                    location=f"office metadata · {field_name}",
                    value=value,
                    rule_id="office-author",
                    remediation=(
                        "Clear creator and last-modified-by properties in the disclosure copy."
                    ),
                )
            )
    return chunks, signals


def _xml_text_chunks(
    archive: zipfile.ZipFile, member_pattern: re.Pattern[str], label: str
) -> list[ContentChunk]:
    chunks: list[ContentChunk] = []
    for member in sorted(archive.namelist()):
        if not member_pattern.fullmatch(member):
            continue
        root = ElementTree.fromstring(archive.read(member))
        text = " ".join(part.strip() for part in root.itertext() if part.strip())
        if text:
            chunks.append(ContentChunk(location=f"{label} · {member}", text=text))
    return chunks


def _format_gps_value(value: Any) -> str:
    if hasattr(value, "items"):
        parts = [
            f"{ExifTags.GPSTAGS.get(int(key), str(key))}={item}" for key, item in value.items()
        ]
        return "; ".join(parts)
    return str(value)


class LocalFileSource:
    """Discover files deterministically without following symlinked directories."""

    def __init__(self, excluded_directories: set[str] | None = None) -> None:
        self.excluded_directories = excluded_directories or DEFAULT_EXCLUDED_DIRECTORIES

    def discover(self, root: Path) -> tuple[Path, ...]:
        if not root.exists():
            raise FileNotFoundError("scan target does not exist")
        if root.is_symlink():
            raise FileAccessError("symlink targets are not scanned")
        if root.is_file():
            return (root,)
        if not root.is_dir():
            raise FileAccessError("scan target is not a regular file or directory")

        files: list[Path] = []
        for current_root, directories, names in os.walk(root, followlinks=False):
            directories[:] = sorted(
                directory
                for directory in directories
                if directory not in self.excluded_directories
                and not (Path(current_root) / directory).is_symlink()
            )
            for name in sorted(names):
                candidate = Path(current_root) / name
                if not candidate.is_symlink() and candidate.is_file():
                    files.append(candidate)
        return tuple(files)

    @staticmethod
    def relative_path(root: Path, candidate: Path) -> str:
        if root.is_file():
            return root.name
        return candidate.relative_to(root).as_posix()


class LocalFileExtractor:
    """Extract supported content without copying source data."""

    def __init__(self, max_file_size: int = 20 * 1024 * 1024) -> None:
        if max_file_size < 1:
            raise ValueError("max_file_size must be positive")
        self.max_file_size = max_file_size

    def extract(self, path: Path) -> ExtractedDocument:
        try:
            size = path.stat().st_size
            if size > self.max_file_size:
                raise FileTooLargeError(
                    f"file exceeds configured {self.max_file_size}-byte safety limit"
                )
            suffix = path.suffix.casefold()
            if suffix not in SUPPORTED_SUFFIXES:
                raise UnsupportedFileError(f"unsupported file type: {suffix or '[none]'}")
            if suffix == ".csv":
                return self._extract_delimited(path, ",")
            if suffix == ".tsv":
                return self._extract_delimited(path, "\t")
            if suffix == ".json":
                return self._extract_json(path)
            if suffix in TEXT_SUFFIXES:
                return self._extract_text(path)
            if suffix == ".xlsx":
                return self._extract_xlsx(path)
            if suffix in {".docx", ".pptx"}:
                return self._extract_office_zip(path, suffix)
            return self._extract_image(path)
        except (UnsupportedFileError, FileTooLargeError):
            raise
        except Exception as error:
            raise FileAccessError(_safe_error(error)) from error

    @staticmethod
    def _read_text(path: Path) -> str:
        raw = path.read_bytes()
        if raw.startswith((b"\xff\xfe", b"\xfe\xff")):
            return raw.decode("utf-16")
        return raw.decode("utf-8-sig")

    def _mime_type(self, path: Path) -> str:
        if magic_module is not None:
            try:
                detected = magic_module.from_file(str(path), mime=True)
                if isinstance(detected, str):
                    return detected
            except (OSError, TypeError):
                pass
        return mimetypes.guess_type(path.name)[0] or "application/octet-stream"

    def _extract_text(self, path: Path) -> ExtractedDocument:
        return ExtractedDocument(
            file_type="text",
            mime_type=self._mime_type(path),
            chunks=(ContentChunk(location="text", text=self._read_text(path)),),
        )

    def _extract_delimited(self, path: Path, delimiter: str) -> ExtractedDocument:
        chunks: list[ContentChunk] = []
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            for row_index, row in enumerate(csv.reader(handle, delimiter=delimiter), start=1):
                for column_index, value in enumerate(row, start=1):
                    if value:
                        chunks.append(
                            ContentChunk(
                                location=f"table · row {row_index}, column {column_index}",
                                text=value,
                            )
                        )
        return ExtractedDocument(
            file_type="csv" if delimiter == "," else "tsv",
            mime_type=self._mime_type(path),
            chunks=tuple(chunks),
        )

    def _extract_json(self, path: Path) -> ExtractedDocument:
        parsed = json.loads(self._read_text(path))
        return ExtractedDocument(
            file_type="json",
            mime_type=self._mime_type(path),
            chunks=tuple(_flatten_json(parsed)),
        )

    def _extract_xlsx(self, path: Path) -> ExtractedDocument:
        chunks: list[ContentChunk] = []
        signals: list[MetadataSignal] = []
        workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
        try:
            for field_name in ("creator", "lastModifiedBy", "title", "subject", "description"):
                value = getattr(workbook.properties, field_name, None)
                if not value:
                    continue
                location = f"office metadata · {field_name}"
                chunks.append(ContentChunk(location=location, text=str(value)))
                if field_name in {"creator", "lastModifiedBy"}:
                    signals.append(
                        MetadataSignal(
                            category=DataCategory.AUTHOR_METADATA,
                            severity=Severity.MEDIUM,
                            location=location,
                            value=str(value),
                            rule_id="office-author",
                            remediation=(
                                "Clear creator and last-modified-by properties in the "
                                "disclosure copy."
                            ),
                        )
                    )
            for worksheet in workbook.worksheets:
                for row in worksheet.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            chunks.append(
                                ContentChunk(
                                    location=f"sheet {worksheet.title} · {cell.coordinate}",
                                    text=str(cell.value),
                                )
                            )
        finally:
            workbook.close()
        return ExtractedDocument(
            file_type="xlsx",
            mime_type=self._mime_type(path),
            chunks=tuple(chunks),
            metadata_signals=tuple(signals),
        )

    def _extract_office_zip(self, path: Path, suffix: str) -> ExtractedDocument:
        with zipfile.ZipFile(path) as archive:
            core_chunks, signals = _office_core_chunks(archive)
            if suffix == ".docx":
                content_chunks = _xml_text_chunks(
                    archive,
                    re.compile(r"word/(?:document|header\d+|footer\d+|comments)\.xml"),
                    "Word content",
                )
            else:
                content_chunks = _xml_text_chunks(
                    archive,
                    re.compile(r"ppt/slides/slide\d+\.xml"),
                    "PowerPoint content",
                )
        return ExtractedDocument(
            file_type=suffix.removeprefix("."),
            mime_type=self._mime_type(path),
            chunks=tuple(core_chunks + content_chunks),
            metadata_signals=tuple(signals),
        )

    def _extract_image(self, path: Path) -> ExtractedDocument:
        chunks: list[ContentChunk] = []
        signals: list[MetadataSignal] = []
        with Image.open(path) as image:
            exif = image.getexif()
            for tag_id, raw_value in exif.items():
                tag = ExifTags.TAGS.get(tag_id, str(tag_id))
                if tag == "GPSInfo":
                    with suppress(KeyError, TypeError):
                        raw_value = exif.get_ifd(tag_id)
                value = _format_gps_value(raw_value)
                if not value:
                    continue
                location = f"EXIF · {tag}"
                chunks.append(ContentChunk(location=location, text=value))
                if tag == "GPSInfo":
                    signals.append(
                        MetadataSignal(
                            category=DataCategory.GPS_LOCATION,
                            severity=Severity.HIGH,
                            location=location,
                            value=value,
                            rule_id="exif-gps",
                            remediation=(
                                "Strip GPS EXIF from the disclosure copy and verify derived images."
                            ),
                        )
                    )
                elif tag in {"Artist", "Copyright"}:
                    signals.append(
                        MetadataSignal(
                            category=DataCategory.AUTHOR_METADATA,
                            severity=Severity.MEDIUM,
                            location=location,
                            value=value,
                            rule_id="exif-author",
                            remediation=(
                                "Remove authorship EXIF when identity is not intended to be public."
                            ),
                        )
                    )
        return ExtractedDocument(
            file_type=f"image/{path.suffix.casefold().removeprefix('.')}",
            mime_type=self._mime_type(path),
            chunks=tuple(chunks),
            metadata_signals=tuple(signals),
        )
